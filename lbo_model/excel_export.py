# =============================================================================
# excel_export.py — Two-sheet Excel: Inputs (blue cells) + Model (formulas)
# =============================================================================
#
# Sheet 1 "Inputs"  — every assumption in a labelled blue cell.
#                     Change any cell here and the whole model recalculates.
# Sheet 2 "Model"   — all sections use Excel formulas referencing Inputs.
#                     Debt sweep uses IF/MIN/MAX.
#                     IRR = MoM^(1/N)−1  (exact for single entry/exit CF).
#
# =============================================================================

from __future__ import annotations
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import inputs as cfg

# ── Colours ───────────────────────────────────────────────────────────────────
BLUE_FONT  = "0000FF"
BLACK_FONT = "000000"
HDR_BG     = "1F4E79"
HDR_FG     = "FFFFFF"
SUB_BG     = "D6E4F0"
HELPER_BG  = "F2F2F2"   # light grey for intermediate debt-sweep helper rows
SENS_HL    = "FFFF00"

# ── Number formats ─────────────────────────────────────────────────────────────
FMT_DOLLAR = '#,##0.0;(#,##0.0);"-"'
FMT_PCT    = '0.0%;(0.0%);"-"'
FMT_MULT   = '0.0"x"'
FMT_IRR    = '0.0%;(0.0%);"-"'
FMT_MOM    = '0.00"x"'
FMT_INT    = '0;(0);"-"'

# ── Column layout (Model sheet) ───────────────────────────────────────────────
COL_LABEL = 1   # A — row labels
COL_TXN   = 2   # B — Year 0 / Transaction column
# Year y is at column  COL_TXN + y  (y=1..N)

LABEL_W = 38
DATA_W  = 13


# =============================================================================
# Helpers
# =============================================================================

def _font(bold=False, color=BLACK_FONT, size=9, italic=False):
    return Font(name="Arial", size=size, bold=bold, color=color, italic=italic)

def _fill(hex_color: str):
    return PatternFill("solid", fgColor=hex_color)

def _align(h="right", wrap=False):
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)

def _top_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(top=s)

def _apply(cell, value=None, fmt=None, bold=False, color=BLACK_FONT,
           fill=None, h="right", italic=False, size=9):
    if value is not None:
        cell.value = value
    cell.font      = _font(bold=bold, color=color, size=size, italic=italic)
    cell.alignment = _align(h)
    if fmt:
        cell.number_format = fmt
    if fill:
        cell.fill = _fill(fill)

def _cl(col_num: int) -> str:
    return get_column_letter(col_num)

def _col(y: int) -> int:
    """Column number for year y (0=Txn/B, 1=C, ...)."""
    return COL_TXN + y

def _ref(row: int, y: int, sheet: str = "") -> str:
    """Absolute cell reference  $C$5  or  'Sheet'!$C$5"""
    addr = f"${_cl(_col(y))}${row}"
    return f"'{sheet}'!{addr}" if sheet else addr

def _single_ref(row: int, col: int, sheet: str = "") -> str:
    addr = f"${_cl(col)}${row}"
    return f"'{sheet}'!{addr}" if sheet else addr


# =============================================================================
# ─────────────────────────────  SHEET 1: INPUTS  ─────────────────────────────
# =============================================================================

# Row numbers for each input cell (column C = value column on Inputs sheet)
INP_COL   = 3   # column C holds the value
INP_SHEET = "Inputs"

_INP_LAYOUT = [
    # (section_header_or_None, constant_name, label, fmt)
    ("ENTRY VALUATION",    None,                     None,                          None),
    (None, "ENTRY_EBITDA",           "Entry EBITDA ($mm)",                  FMT_DOLLAR),
    (None, "ENTRY_MULTIPLE",         "(x) LTM Entry Multiple",              FMT_MULT),
    (None, "EXISTING_DEBT",          "Existing Target Debt ($mm)",          FMT_DOLLAR),
    (None, "EXISTING_CASH",          "Existing Target Cash ($mm)",          FMT_DOLLAR),
    ("TRANSACTION",        None,                     None,                          None),
    (None, "EXIT_MULTIPLE",          "LTM Exit Multiple",                   FMT_MULT),
    (None, "TAX_RATE",               "Tax Rate",                            FMT_PCT),
    (None, "CASH_TO_BS",             "Cash to Balance Sheet ($mm)",         FMT_DOLLAR),
    (None, "TRANSACTION_EXPENSES",   "Transaction Expenses ($mm)",          FMT_DOLLAR),
    (None, "ROLLOVER_EQUITY_PCT",    "Rollover Equity (%)",                 FMT_PCT),
    ("DEBT — GENERAL",     None,                     None,                          None),
    (None, "BASE_RATE",              "Flat Base Rate / SOFR",               FMT_PCT),
    (None, "LEVERAGEABLE_EBITDA",    "Leverageable EBITDA ($mm)",           FMT_DOLLAR),
    (None, "MIN_CASH_BALANCE",       "Minimum Cash Balance ($mm)",          FMT_DOLLAR),
    (None, "FINANCING_FEE_AMORT_YEARS", "Financing Fee Amort (years)",      FMT_INT),
    ("DEBT — TERM LOAN",   None,                     None,                          None),
    (None, "TERM_LOAN_LEVERAGE",     "Leverage (x EBITDA)",                 FMT_MULT),
    (None, "TERM_LOAN_SPREAD",       "Spread (L+)",                         FMT_PCT),
    (None, "TERM_LOAN_FEE_PCT",      "Upfront Fee (%)",                     FMT_PCT),
    (None, "TERM_LOAN_AMORT_RATE",   "Mandatory Amort (% of original)",     FMT_PCT),
    ("DEBT — MEZZANINE",   None,                     None,                          None),
    (None, "MEZZ_LEVERAGE",          "Leverage (x EBITDA)",                 FMT_MULT),
    (None, "MEZZ_CASH_RATE",         "Cash Pay Rate",                       FMT_PCT),
    (None, "MEZZ_PIK_RATE",          "PIK Rate",                            FMT_PCT),
    (None, "MEZZ_FEE_PCT",           "Upfront Fee (%)",                     FMT_PCT),
    (None, "MEZZ_CASH_SWEEP",        "Sweep Mezz After TL Repaid? (1=Y)",   FMT_INT),
    ("DEBT — REVOLVER",    None,                     None,                          None),
    (None, "REVOLVER_CAPACITY",      "Revolver Capacity ($mm)",             FMT_DOLLAR),
    (None, "REVOLVER_COMMITMENT_FEE","Commitment Fee (% of undrawn)",       FMT_PCT),
    (None, "REVOLVER_SPREAD",        "Spread (L+, when drawn)",             FMT_PCT),
    ("FINANCIAL FORECAST", None,                     None,                          None),
    (None, "BASE_REVENUE",           "Base Revenue — Year 0 ($mm)",         FMT_DOLLAR),
    (None, "REVENUE_GROWTH_RATE",    "Annual Revenue Growth Rate",          FMT_PCT),
    (None, "GROSS_MARGIN",           "Gross Margin (% of Revenue)",         FMT_PCT),
    (None, "SGA_PCT",                "SG&A (% of Revenue)",                 FMT_PCT),
    (None, "RD_PCT",                 "R&D (% of Revenue)",                  FMT_PCT),
    (None, "DA_PCT",                 "D&A (% of Revenue)",                  FMT_PCT),
    (None, "CAPEX_PCT",              "Capex (% of Revenue)",                FMT_PCT),
    (None, "NWC_PCT",                "Increase in NWC (% of Revenue)",      FMT_PCT),
    ("HOLD PERIOD",        None,                     None,                          None),
    (None, "HOLD_YEARS",             "Hold Period (years)",                  FMT_INT),
]


def _write_inputs_sheet(wb: Workbook, overrides: dict) -> dict:
    """
    Write the Inputs sheet. Returns a dict {constant_name: abs_cell_ref_string}
    so the Model sheet can build formulas referencing these cells.
    """
    ws = wb.create_sheet(INP_SHEET)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 18

    # Title
    t = ws.cell(1, 2, "LBO Model Template — Inputs")
    _apply(t, bold=True, size=13, color=HDR_BG, h="left")
    t2 = ws.cell(2, 2, "Change any blue cell — the Model sheet recalculates automatically.")
    _apply(t2, size=9, color="595959", italic=True, h="left")
    ws.cell(3, 2, "($ in millions unless labelled)")
    _apply(ws.cell(3, 2), size=9, color="595959", italic=True, h="left")

    row = 5
    inp_cells: dict[str, str] = {}

    for section, key, label, fmt in _INP_LAYOUT:
        if section is not None:
            # Section header
            row += 1
            c = ws.cell(row, 2, section)
            _apply(c, bold=True, color=HDR_FG, fill=HDR_BG, h="left", size=9)
            ws.merge_cells(start_row=row, start_column=2,
                           end_row=row,   end_column=3)
            row += 1
        else:
            # Data row
            c_label = ws.cell(row, 2, "  " + label)
            _apply(c_label, h="left", size=9)

            # Value: use override if provided, else module default
            raw_val = overrides.get(key, getattr(cfg, key))
            # MEZZ_CASH_SWEEP: store as 1/0 in Excel
            if key == "MEZZ_CASH_SWEEP":
                raw_val = 1 if raw_val else 0

            c_val = ws.cell(row, INP_COL, raw_val)
            _apply(c_val, fmt=fmt, color=BLUE_FONT, size=9)

            # Record absolute reference
            inp_cells[key] = f"'{INP_SHEET}'!${_cl(INP_COL)}${row}"
            row += 1

    return inp_cells


# =============================================================================
# ─────────────────────────────  SHEET 2: MODEL  ──────────────────────────────
# =============================================================================

class _W:
    """Row-tracking writer for the Model sheet.

    Two-pass simulation support:
        simulate=True  → rows are counted and reg[] is populated, but nothing
                         is written to the worksheet.  r() returns a dummy ref
                         so forward-reference list comprehensions don't crash.
        simulate=False → normal write mode; reg[] is pre-populated from the
                         simulation pass so every r() call resolves correctly.
    """
    def __init__(self, ws, N: int, simulate: bool = False):
        self.ws       = ws
        self.row      = 1
        self.N        = N
        self.simulate = simulate
        self.reg: dict[str, int] = {}   # name → row number

    # ── Row registration ──────────────────────────────────────────────────────
    def track(self, name: str):
        self.reg[name] = self.row

    def ref(self, name: str, y: int) -> str:
        """Absolute ref to a tracked model row at year y.
        In simulate mode returns a safe dummy so callers don't raise KeyError."""
        if self.simulate:
            return "$A$1"
        return _ref(self.reg[name], y)

    def txn_ref(self, name: str) -> str:
        return self.ref(name, 0)

    # ── Layout helpers ────────────────────────────────────────────────────────
    def blank(self, n=1):
        self.row += n

    def section_header(self, text: str):
        end = _col(self.N)
        if not self.simulate:
            c = self.ws.cell(self.row, COL_LABEL, text)
            _apply(c, bold=True, color=HDR_FG, fill=HDR_BG, h="left", size=10)
            self.ws.merge_cells(start_row=self.row, start_column=COL_LABEL,
                                end_row=self.row,   end_column=end)
        self.row += 1

    def sub_header(self, text: str):
        end = _col(self.N)
        if not self.simulate:
            c = self.ws.cell(self.row, COL_LABEL, text)
            _apply(c, bold=True, fill=SUB_BG, h="left", size=9)
            self.ws.merge_cells(start_row=self.row, start_column=COL_LABEL,
                                end_row=self.row,   end_column=end)
        self.row += 1

    def year_header_row(self, txn_label="Txn"):
        if not self.simulate:
            r = self.row
            _apply(self.ws.cell(r, COL_LABEL, ""), bold=True, fill=SUB_BG, h="left", size=9)
            _apply(self.ws.cell(r, _col(0), txn_label), bold=True, fill=SUB_BG, size=9)
            for y in range(1, self.N + 1):
                _apply(self.ws.cell(r, _col(y), f"Year {y}"), bold=True, fill=SUB_BG, size=9)
        self.row += 1

    # ── Data writers ─────────────────────────────────────────────────────────

    def formula_row(self, name: str | None, label: str,
                    formulas: list,          # length N+1: [txn_formula, y1, y2, ...]
                    fmt=FMT_DOLLAR, bold=False, color=BLACK_FONT,
                    helper=False, indent=2):
        """
        Write one row of formulas.
        formulas[0] = txn/Year-0 column; formulas[1..N] = Year 1..N.
        None in formulas skips that cell.
        """
        r = self.row
        if name:
            self.reg[name] = r
        if not self.simulate:
            bg = HELPER_BG if helper else None
            lbl_cell = self.ws.cell(r, COL_LABEL, " " * indent + label)
            _apply(lbl_cell, h="left", bold=bold, size=9, fill=bg)
            for y, fml in enumerate(formulas):
                if fml is None:
                    continue
                c = self.ws.cell(r, _col(y), fml)
                _apply(c, fmt=fmt, bold=bold, color=color, size=9, fill=bg)
        self.row += 1

    def total_row(self, name: str | None, label: str,
                  formulas: list, fmt=FMT_DOLLAR):
        """Bold row with top border."""
        r = self.row
        if name:
            self.reg[name] = r
        if not self.simulate:
            lbl = self.ws.cell(r, COL_LABEL, "  " + label)
            _apply(lbl, bold=True, h="left", size=9)
            for y, fml in enumerate(formulas):
                if fml is None:
                    continue
                c = self.ws.cell(r, _col(y), fml)
                _apply(c, fmt=fmt, bold=True, size=9)
                c.border = _top_border()
        self.row += 1

    def value_row(self, name: str | None, label: str,
                  values: list, fmt=FMT_DOLLAR, bold=False,
                  color=BLACK_FONT, indent=2):
        """Write Python-computed values (not formulas) — for sensitivity table axis labels etc."""
        r = self.row
        if name:
            self.reg[name] = r
        if not self.simulate:
            _apply(self.ws.cell(r, COL_LABEL, " " * indent + label), h="left", bold=bold, size=9)
            for col_idx, val in enumerate(values):
                if val is None:
                    continue
                c = self.ws.cell(r, col_idx, val)
                _apply(c, fmt=fmt, bold=bold, color=color, size=9)
        self.row += 1


# =============================================================================
# Formula builders  (pure strings — no Python arithmetic)
# =============================================================================

def _build_model_formulas(W: _W, I: dict, N: int):
    """
    Write every section to the Model sheet using Excel formula strings.
    I = inp_cells dict from _write_inputs_sheet()
    W = _W writer
    """
    ws = W.ws

    # Convenience aliases
    def i(key):
        return I[key]  # e.g. "'Inputs'!$C$5"

    def r(name, y):
        return W.ref(name, y)  # e.g. "$C$12"

    # =========================================================================
    # TITLE
    # =========================================================================
    if not W.simulate:
        t = ws.cell(W.row, COL_LABEL, "LBO Model Template")
        _apply(t, bold=True, size=14, color=HDR_BG, h="left")
    W.row += 1
    if not W.simulate:
        t2 = ws.cell(W.row, COL_LABEL, "($ in millions)   |   Change assumptions on the Inputs sheet")
        _apply(t2, size=9, color="595959", italic=True, h="left")
    W.blank(2)

    # =========================================================================
    # 1. TRANSACTION DRIVERS
    # =========================================================================
    W.section_header("1.  TRANSACTION DRIVERS")
    W.sub_header("Entry Valuation")

    # These rows cascade — each references the previous one.
    W.formula_row("purchase_ev",   "Purchase Enterprise Value",
                  [f"={i('ENTRY_EBITDA')}*{i('ENTRY_MULTIPLE')}", None,None,None,None,None],
                  FMT_DOLLAR, indent=2)
    W.formula_row(None, "Entry EBITDA ($mm)",
                  [f"={i('ENTRY_EBITDA')}", None,None,None,None,None],
                  FMT_DOLLAR, color=BLUE_FONT, indent=4)
    W.formula_row(None, "(x) LTM Entry Multiple",
                  [f"={i('ENTRY_MULTIPLE')}", None,None,None,None,None],
                  FMT_MULT, color=BLUE_FONT, indent=4)
    W.formula_row(None, "Less: Existing Debt",
                  [f"=-{i('EXISTING_DEBT')}", None,None,None,None,None],
                  FMT_DOLLAR, indent=4)
    W.formula_row(None, "Plus: Existing Cash",
                  [f"={i('EXISTING_CASH')}", None,None,None,None,None],
                  FMT_DOLLAR, indent=4)
    W.formula_row("equity_value", "Equity Value",
                  [f"={W.txn_ref('purchase_ev')}-{i('EXISTING_DEBT')}+{i('EXISTING_CASH')}",
                   None,None,None,None,None],
                  FMT_DOLLAR, bold=True, indent=2)
    W.blank()

    W.sub_header("Transaction Assumptions")
    W.formula_row(None,"LTM Exit Multiple",
                  [f"={i('EXIT_MULTIPLE')}",None,None,None,None,None], FMT_MULT, color=BLUE_FONT)
    W.formula_row(None,"Tax Rate",
                  [f"={i('TAX_RATE')}",None,None,None,None,None], FMT_PCT, color=BLUE_FONT)
    W.formula_row(None,"Cash to B/S ($mm)",
                  [f"={i('CASH_TO_BS')}",None,None,None,None,None], FMT_DOLLAR, color=BLUE_FONT)
    W.formula_row(None,"Rollover Equity (%)",
                  [f"={i('ROLLOVER_EQUITY_PCT')}",None,None,None,None,None], FMT_PCT, color=BLUE_FONT)
    W.formula_row("pe_ownership","PE Ownership (%)",
                  [f"=1-{i('ROLLOVER_EQUITY_PCT')}",None,None,None,None,None], FMT_PCT)
    W.blank()

    W.sub_header("Debt Sizing")
    W.formula_row("tl_principal","Term Loan Principal ($mm)",
                  [f"={i('LEVERAGEABLE_EBITDA')}*{i('TERM_LOAN_LEVERAGE')}",None,None,None,None,None],
                  FMT_DOLLAR)
    W.formula_row("mezz_principal","Mezzanine Principal ($mm)",
                  [f"={i('LEVERAGEABLE_EBITDA')}*{i('MEZZ_LEVERAGE')}",None,None,None,None,None],
                  FMT_DOLLAR)
    W.formula_row("total_debt","Total Debt ($mm)",
                  [f"={W.txn_ref('tl_principal')}+{W.txn_ref('mezz_principal')}",None,None,None,None,None],
                  FMT_DOLLAR, bold=True)
    W.formula_row("tl_fee","TL Upfront Fee ($mm)",
                  [f"={W.txn_ref('tl_principal')}*{i('TERM_LOAN_FEE_PCT')}",None,None,None,None,None],
                  FMT_DOLLAR)
    W.formula_row("mezz_fee","Mezz Upfront Fee ($mm)",
                  [f"={W.txn_ref('mezz_principal')}*{i('MEZZ_FEE_PCT')}",None,None,None,None,None],
                  FMT_DOLLAR)
    W.formula_row("total_fees","Total Financing Fees ($mm)",
                  [f"={W.txn_ref('tl_fee')}+{W.txn_ref('mezz_fee')}",None,None,None,None,None],
                  FMT_DOLLAR)
    W.formula_row("annual_fee_amort","Annual Fee Amortization ($mm)",
                  [f"=IF({i('FINANCING_FEE_AMORT_YEARS')}>0,{W.txn_ref('total_fees')}/{i('FINANCING_FEE_AMORT_YEARS')},0)",
                   None,None,None,None,None], FMT_DOLLAR)
    W.blank()

    W.sub_header("Sources & Uses")
    # Uses
    W.formula_row("refinancing_nd","Refinancing Net Debt",
                  [f"={i('EXISTING_DEBT')}-{i('EXISTING_CASH')}",None,None,None,None,None], FMT_DOLLAR)
    W.formula_row("total_uses","Total Uses",
                  [f"={W.txn_ref('purchase_ev')}+{i('CASH_TO_BS')}+{W.txn_ref('total_fees')}+{i('TRANSACTION_EXPENSES')}",
                   None,None,None,None,None], FMT_DOLLAR, bold=True)
    # Sources
    W.formula_row("total_equity","Total Equity",
                  [f"={W.txn_ref('total_uses')}-{W.txn_ref('total_debt')}",None,None,None,None,None],
                  FMT_DOLLAR)
    W.formula_row("rollover_equity","Rollover Equity",
                  [f"={W.txn_ref('total_equity')}*{i('ROLLOVER_EQUITY_PCT')}",None,None,None,None,None],
                  FMT_DOLLAR)
    W.formula_row("pe_invest","PE Equity Investment",
                  [f"={W.txn_ref('total_equity')}*(1-{i('ROLLOVER_EQUITY_PCT')})",None,None,None,None,None],
                  FMT_DOLLAR, bold=True)
    W.formula_row(None,"Total Sources",
                  [f"={W.txn_ref('total_debt')}+{W.txn_ref('total_equity')}",None,None,None,None,None],
                  FMT_DOLLAR, bold=True)
    W.formula_row(None,"  Check (Sources − Uses)",
                  [f"={W.txn_ref('total_debt')}+{W.txn_ref('total_equity')}-{W.txn_ref('total_uses')}",
                   None,None,None,None,None], FMT_DOLLAR)
    W.blank()

    # =========================================================================
    # 2. FINANCIAL FORECAST
    # =========================================================================
    W.section_header("2.  FINANCIAL FORECAST")
    W.year_header_row("Year 0")

    # Revenue
    rev_fmls = [f"={i('BASE_REVENUE')}"] + [
        f"={r('revenue', y-1)}*(1+{i('REVENUE_GROWTH_RATE')})" for y in range(1, N+1)
    ]
    W.formula_row("revenue", "Revenue", rev_fmls, FMT_DOLLAR)

    growth_fmls = [None] + [
        f"={r('revenue',y)}/{r('revenue',y-1)}-1" for y in range(1, N+1)
    ]
    W.formula_row(None, "% Growth", growth_fmls, FMT_PCT, indent=4)
    W.blank()

    gp_fmls = [None] + [f"={r('revenue',y)}*{i('GROSS_MARGIN')}" for y in range(1,N+1)]
    W.formula_row("gross_profit", "Gross Profit", gp_fmls, FMT_DOLLAR)
    gpm = [None] + [f"={r('gross_profit',y)}/{r('revenue',y)}" for y in range(1,N+1)]
    W.formula_row(None, "% Margin", gpm, FMT_PCT, indent=4)
    W.blank()

    sga_fmls = [None] + [f"=-{r('revenue',y)}*{i('SGA_PCT')}" for y in range(1,N+1)]
    rd_fmls  = [None] + [f"=-{r('revenue',y)}*{i('RD_PCT')}"  for y in range(1,N+1)]
    W.formula_row("sga", "Less: SG&A", sga_fmls, FMT_DOLLAR)
    W.formula_row("rd",  "Less: R&D",  rd_fmls,  FMT_DOLLAR)

    ebitda_txn  = f"={i('ENTRY_EBITDA')}"
    ebitda_fmls = [ebitda_txn] + [
        f"={r('gross_profit',y)}+{r('sga',y)}+{r('rd',y)}" for y in range(1,N+1)
    ]
    W.total_row("ebitda", "EBITDA", ebitda_fmls)
    margin_fmls = [f"={W.txn_ref('ebitda')}/{i('BASE_REVENUE')}"] + [
        f"={r('ebitda',y)}/{r('revenue',y)}" for y in range(1,N+1)
    ]
    W.formula_row(None, "% Margin", margin_fmls, FMT_PCT, indent=4)
    W.blank()

    da_fmls   = [None] + [f"=-{r('revenue',y)}*{i('DA_PCT')}" for y in range(1,N+1)]
    W.formula_row("da", "Less: D&A", da_fmls, FMT_DOLLAR)
    ebit_fmls = [None] + [f"={r('ebitda',y)}+{r('da',y)}" for y in range(1,N+1)]
    W.total_row("ebit", "EBIT", ebit_fmls)
    W.blank()

    # Interest — references debt balances (built in debt schedule below).
    # We write these rows now but the formula references point FORWARD to the
    # debt schedule rows.  In Excel, forward references work fine.
    # We need to pre-register the debt balance row names so we can reference them.
    # Strategy: declare placeholder row numbers here, then write the debt rows later.
    # We'll reserve names using a forward-declaration trick — write the interest
    # rows with a reference to names that WILL exist when debt schedule is built.
    #
    # Simpler approach: build debt schedule rows FIRST, then come back to interest.
    # We achieve this by splitting the write into two passes — but that's complex.
    #
    # CLEANEST: write the sections in the natural order (forecast first, then debt),
    # using INDIRECT or just accepting that the formulas reference rows further down.
    # Excel resolves forward references at calc time; openpyxl writes them as strings.
    # We just need the row numbers in advance.  Solution: pre-compute the row offset.
    #
    # We'll estimate where debt schedule rows will land by counting rows from current.
    # Instead, let's use a two-pass approach: record the CURRENT row, skip the
    # interest/EBT/NI/FCF rows, write the debt schedule, then come back and fill in.
    #
    # PRACTICAL APPROACH: write debt BEFORE income statement continuation.
    # Reorder the sections: Rev→GP→EBITDA→EBIT, then DEBT SCHEDULE, then resume IS.
    # This is actually cleaner from an auditor's perspective too.
    # We'll do: Revenue section → Debt Schedule → Interest section → EBT→NI→FCF

    # -- skip to debt schedule, then come back --
    # Reserve row slots for the interest block:
    int_block_start = W.row   # remember where we are
    W.blank(18)               # skip 18 rows (placeholder for interest+EBT+NI+FCF bridge)
    IS_PLACEHOLDER = int_block_start

    # =========================================================================
    # 3. DEBT SCHEDULE
    # =========================================================================
    W.section_header("3.  DEBT SCHEDULE")

    # ── Revolver ──────────────────────────────────────────────────────────────
    W.sub_header("Revolver")
    W.year_header_row("Txn")

    # Beginning balance: Y1 = 0, Yy = rev_end[y-1]
    rev_beg_fmls = [f"=0"] + [
        "=0" if y == 1 else f"={r('rev_end', y-1)}" for y in range(1, N+1)
    ]
    W.formula_row("rev_beg", "Beginning Balance", rev_beg_fmls, FMT_DOLLAR)

    # Cash pool = beg_cash + FCF_before_optional  (beg_cash = cash_end[y-1]; FCF_before registered later)
    # We reference 'cash_end' and 'fcf_before_opt' — both registered later.
    # This is a forward reference — fine in Excel.
    cash_pool_fmls = [None] + [
        f"={r('cash_end', y-1)}+{r('fcf_before_opt', y)}" for y in range(1, N+1)
    ]
    W.formula_row("cash_pool", "  Cash Pool (Beg Cash + FCF-Bef-Opt)", cash_pool_fmls, FMT_DOLLAR, helper=True)

    rev_draw_fmls = [None] + [
        f"=MAX(0,MIN({i('MIN_CASH_BALANCE')}-{r('cash_pool',y)},{i('REVOLVER_CAPACITY')}-{r('rev_beg',y)}))"
        for y in range(1, N+1)
    ]
    W.formula_row("rev_draw", "  Revolver Draw", rev_draw_fmls, FMT_DOLLAR, helper=True)

    rev_repay_fmls = [None] + [
        f"=IF({r('cash_pool',y)}>={i('MIN_CASH_BALANCE')},MIN({r('cash_pool',y)}-{i('MIN_CASH_BALANCE')},{r('rev_beg',y)}),0)"
        for y in range(1, N+1)
    ]
    W.formula_row("rev_repay", "  (Repayment)", rev_repay_fmls, FMT_DOLLAR, helper=True)

    rev_end_fmls = [f"=0"] + [
        f"={r('rev_beg',y)}+{r('rev_draw',y)}-{r('rev_repay',y)}" for y in range(1,N+1)
    ]
    W.total_row("rev_end", "Ending Balance", rev_end_fmls)

    cash_after_rev_fmls = [None] + [
        f"={r('cash_pool',y)}+{r('rev_draw',y)}-{r('rev_repay',y)}" for y in range(1,N+1)
    ]
    W.formula_row("cash_after_rev","  Cash After Revolver", cash_after_rev_fmls, FMT_DOLLAR, helper=True)

    rev_int_fmls = [None] + [
        f"={r('rev_beg',y)}*({i('BASE_RATE')}+{i('REVOLVER_SPREAD')})+({i('REVOLVER_CAPACITY')}-{r('rev_beg',y)})*{i('REVOLVER_COMMITMENT_FEE')}"
        for y in range(1,N+1)
    ]
    W.formula_row("rev_interest","Interest & Commitment Fee", rev_int_fmls, FMT_DOLLAR)
    W.blank()

    # ── Term Loan ─────────────────────────────────────────────────────────────
    W.sub_header("Term Loan")
    W.year_header_row("Txn")

    tl_beg_fmls = [f"={W.txn_ref('tl_principal')}"] + [
        f"={W.txn_ref('tl_principal')}" if y == 1 else f"={r('tl_end',y-1)}"
        for y in range(1, N+1)
    ]
    W.formula_row("tl_beg","Beginning Balance", tl_beg_fmls, FMT_DOLLAR)

    mand_fmls = [None] + [
        f"=MIN({W.txn_ref('tl_principal')}*{i('TERM_LOAN_AMORT_RATE')},{r('tl_beg',y)})"
        for y in range(1,N+1)
    ]
    W.formula_row("mand_amort","  (Mandatory Amortization)", mand_fmls, FMT_DOLLAR)

    tl_after_mand_fmls = [None] + [
        f"=MAX(0,{r('tl_beg',y)}-{r('mand_amort',y)})" for y in range(1,N+1)
    ]
    W.formula_row("tl_after_mand","  Balance After Mandatory Amort", tl_after_mand_fmls, FMT_DOLLAR, helper=True)

    tl_opt_fmls = [None] + [
        f"=MIN(MAX(0,{r('cash_after_rev',y)}-{i('MIN_CASH_BALANCE')}),{r('tl_after_mand',y)})"
        for y in range(1,N+1)
    ]
    W.formula_row("tl_opt","  (Discretionary Repayment)", tl_opt_fmls, FMT_DOLLAR)

    tl_end_fmls = [f"={W.txn_ref('tl_principal')}"] + [
        f"={r('tl_after_mand',y)}-{r('tl_opt',y)}" for y in range(1,N+1)
    ]
    W.total_row("tl_end","Ending Balance", tl_end_fmls)

    cash_after_tl_fmls = [None] + [
        f"={r('cash_after_rev',y)}-{r('tl_opt',y)}" for y in range(1,N+1)
    ]
    W.formula_row("cash_after_tl","  Cash After TL Sweep", cash_after_tl_fmls, FMT_DOLLAR, helper=True)

    tl_rate_fmls = [None] + [f"={i('BASE_RATE')}+{i('TERM_LOAN_SPREAD')}"]*N
    W.formula_row(None,"All-in Rate", tl_rate_fmls, FMT_PCT, color=BLUE_FONT)

    tl_int_fmls = [None] + [
        f"={r('tl_beg',y)}*({i('BASE_RATE')}+{i('TERM_LOAN_SPREAD')})" for y in range(1,N+1)
    ]
    W.formula_row("tl_interest","Interest Expense", tl_int_fmls, FMT_DOLLAR)
    W.blank()

    # ── Mezzanine ─────────────────────────────────────────────────────────────
    W.sub_header("Mezzanine (PIK) Debt")
    W.year_header_row("Txn")

    mezz_beg_fmls = [f"={W.txn_ref('mezz_principal')}"] + [
        f"={W.txn_ref('mezz_principal')}" if y == 1 else f"={r('mezz_end',y-1)}"
        for y in range(1, N+1)
    ]
    W.formula_row("mezz_beg","Beginning Balance", mezz_beg_fmls, FMT_DOLLAR)

    pik_fmls = [None] + [
        f"={r('mezz_beg',y)}*{i('MEZZ_PIK_RATE')}" for y in range(1,N+1)
    ]
    W.formula_row("mezz_pik","  Plus: PIK Accretion", pik_fmls, FMT_DOLLAR)

    mezz_after_pik_fmls = [None] + [
        f"={r('mezz_beg',y)}+{r('mezz_pik',y)}" for y in range(1,N+1)
    ]
    W.formula_row("mezz_after_pik","  Balance After PIK", mezz_after_pik_fmls, FMT_DOLLAR, helper=True)

    # Mezz sweep: only if MEZZ_CASH_SWEEP toggle (Inputs cell) = 1
    mezz_opt_fmls = [None] + [
        f"=IF({i('MEZZ_CASH_SWEEP')}=1,MIN(MAX(0,{r('cash_after_tl',y)}-{i('MIN_CASH_BALANCE')}),{r('mezz_after_pik',y)}),0)"
        for y in range(1,N+1)
    ]
    W.formula_row("mezz_opt","  (Discretionary Repayment)", mezz_opt_fmls, FMT_DOLLAR)

    mezz_end_fmls = [f"={W.txn_ref('mezz_principal')}"] + [
        f"={r('mezz_after_pik',y)}-{r('mezz_opt',y)}" for y in range(1,N+1)
    ]
    W.total_row("mezz_end","Ending Balance", mezz_end_fmls)

    mezz_cash_int_fmls = [None] + [
        f"={r('mezz_beg',y)}*{i('MEZZ_CASH_RATE')}" for y in range(1,N+1)
    ]
    W.formula_row("mezz_cash_int","Cash Interest Expense", mezz_cash_int_fmls, FMT_DOLLAR)
    W.blank()

    # ── Cash ending balance ───────────────────────────────────────────────────
    # cash_end[0] = CASH_TO_BS (txn column)
    # cash_end[y] = cash_after_tl[y] - mezz_opt[y]
    cash_end_fmls = [f"={i('CASH_TO_BS')}"] + [
        f"={r('cash_after_tl',y)}-{r('mezz_opt',y)}" for y in range(1,N+1)
    ]
    W.formula_row("cash_end","Cash Balance (End of Year)", cash_end_fmls, FMT_DOLLAR)
    W.blank()

    # =========================================================================
    # NOW GO BACK AND FILL IN THE INCOME STATEMENT / FCF ROWS
    # =========================================================================
    # We paused at IS_PLACEHOLDER. Now that we've registered all debt rows,
    # we can write those formulas in the reserved rows.
    # We'll use direct ws.cell() writes (bypassing the row counter).

    irow = IS_PLACEHOLDER   # current insert row

    def _fill_row(name, label, fmls, fmt=FMT_DOLLAR, bold=False,
                  color=BLACK_FONT, total=False, helper=False, indent=2):
        nonlocal irow
        if name:
            W.reg[name] = irow
        if not W.simulate:
            bg = HELPER_BG if helper else None
            lbl = ws.cell(irow, COL_LABEL, " "*indent + label)
            _apply(lbl, h="left", bold=bold, size=9, fill=bg)
            for y, fml in enumerate(fmls):
                if fml is None:
                    continue
                c = ws.cell(irow, _col(y), fml)
                _apply(c, fmt=fmt, bold=bold, color=color, size=9, fill=bg)
                if total:
                    c.border = _top_border()
        irow += 1

    total_cash_int_fmls = [None] + [
        f"={r('rev_interest',y)}+{r('tl_interest',y)}+{r('mezz_cash_int',y)}"
        for y in range(1,N+1)
    ]
    _fill_row("total_cash_int","Less: Total Cash Interest",
              [None]+[f"=-({r('rev_interest',y)}+{r('tl_interest',y)}+{r('mezz_cash_int',y)})" for y in range(1,N+1)],
              FMT_DOLLAR)

    pik_is_fmls = [None] + [f"=-{r('mezz_pik',y)}" for y in range(1,N+1)]
    _fill_row("pik_is","Less: PIK Interest (non-cash)", pik_is_fmls, FMT_DOLLAR)

    fee_fmls_raw = [None] + [
        f"=IF({y}<={i('FINANCING_FEE_AMORT_YEARS')},{W.txn_ref('annual_fee_amort')},0)"
        for y in range(1,N+1)
    ]
    # Store fee_amort row (positive value used in FCF addback)
    fee_amort_fmls_pos = [None] + [
        f"=IF({y}<={i('FINANCING_FEE_AMORT_YEARS')},{W.txn_ref('annual_fee_amort')},0)"
        for y in range(1,N+1)
    ]
    fee_amort_fmls_neg = [None] + [
        f"=-IF({y}<={i('FINANCING_FEE_AMORT_YEARS')},{W.txn_ref('annual_fee_amort')},0)"
        for y in range(1,N+1)
    ]
    _fill_row("fee_amort","Less: Amort of Financing Fees", fee_amort_fmls_neg, FMT_DOLLAR)

    ebt_fmls = [None] + [
        f"={r('ebit',y)}+{r('total_cash_int',y)}+{r('pik_is',y)}+{r('fee_amort',y)}"
        for y in range(1,N+1)
    ]
    # total_cash_int, pik_is, fee_amort rows hold NEGATIVE values → adding them subtracts
    _fill_row("ebt","EBT", [None]+[
        f"={r('ebit',y)}-({r('rev_interest',y)}+{r('tl_interest',y)}+{r('mezz_cash_int',y)})-{r('mezz_pik',y)}-IF({y}<={i('FINANCING_FEE_AMORT_YEARS')},{W.txn_ref('annual_fee_amort')},0)"
        for y in range(1,N+1)
    ], FMT_DOLLAR, total=True)

    taxes_fmls = [None] + [
        f"=-MAX(0,{r('ebt',y)})*{i('TAX_RATE')}" for y in range(1,N+1)
    ]
    _fill_row(None,"Less: Taxes", taxes_fmls, FMT_DOLLAR)

    ni_fmls = [None] + [f"={r('ebt',y)}+({r('ebt',y)}>0)*(-{r('ebt',y)}*{i('TAX_RATE')})" for y in range(1,N+1)]
    # Simpler: NI = EBT - taxes = EBT - MAX(0,EBT)*TAX
    ni_fmls2 = [None] + [
        f"={r('ebt',y)}-MAX(0,{r('ebt',y)})*{i('TAX_RATE')}" for y in range(1,N+1)
    ]
    _fill_row("net_income","Net Income", ni_fmls2, FMT_DOLLAR, total=True)

    irow += 1  # blank

    # FCF Bridge header (write directly)
    if not W.simulate:
        c = ws.cell(irow, COL_LABEL, "  Free Cash Flow Bridge")
        _apply(c, bold=True, fill=SUB_BG, h="left", size=9)
        ws.merge_cells(start_row=irow, start_column=COL_LABEL,
                       end_row=irow,   end_column=_col(N))
    irow += 1

    _fill_row(None,"Net Income",
              [None]+[f"={r('net_income',y)}" for y in range(1,N+1)], FMT_DOLLAR)
    _fill_row(None,"Plus: D&A",
              [None]+[f"=-{r('da',y)}" for y in range(1,N+1)], FMT_DOLLAR)
    # da rows store NEGATIVE values, negate them back to positive
    _fill_row(None,"Plus: D&A (addback)",
              [None]+[f"={r('revenue',y)}*{i('DA_PCT')}" for y in range(1,N+1)],
              FMT_DOLLAR, indent=4)
    irow -= 1  # reuse the last row (write over previous duplicate)
    _fill_row(None,"Plus: D&A",
              [None]+[f"={r('revenue',y)}*{i('DA_PCT')}" for y in range(1,N+1)], FMT_DOLLAR)

    _fill_row(None,"Plus: PIK Interest (non-cash)",
              [None]+[f"={r('mezz_pik',y)}" for y in range(1,N+1)], FMT_DOLLAR)
    _fill_row(None,"Plus: Amort of Financing Fees",
              [None]+[
                  f"=IF({y}<={i('FINANCING_FEE_AMORT_YEARS')},{W.txn_ref('annual_fee_amort')},0)"
                  for y in range(1,N+1)
              ], FMT_DOLLAR)
    _fill_row(None,"Less: Capex",
              [None]+[f"=-{r('revenue',y)}*{i('CAPEX_PCT')}" for y in range(1,N+1)], FMT_DOLLAR)
    _fill_row(None,"Less: Increase in NWC",
              [None]+[f"=-{r('revenue',y)}*{i('NWC_PCT')}" for y in range(1,N+1)], FMT_DOLLAR)

    fcf_fmls = [None] + [
        f"={r('net_income',y)}"
        f"+{r('revenue',y)}*{i('DA_PCT')}"
        f"+{r('mezz_pik',y)}"
        f"+IF({y}<={i('FINANCING_FEE_AMORT_YEARS')},{W.txn_ref('annual_fee_amort')},0)"
        f"-{r('revenue',y)}*{i('CAPEX_PCT')}"
        f"-{r('revenue',y)}*{i('NWC_PCT')}"
        for y in range(1,N+1)
    ]
    _fill_row("fcf","Free Cash Flow", fcf_fmls, FMT_DOLLAR, bold=True, total=True)

    _fill_row("mand_amort_is","Less: Mandatory Amortization",
              [None]+[f"=-{r('mand_amort',y)}" for y in range(1,N+1)], FMT_DOLLAR)

    fcf_before_fmls = [None] + [
        f"={r('fcf',y)}-{r('mand_amort',y)}" for y in range(1,N+1)
    ]
    _fill_row("fcf_before_opt","FCF Before Optional Paydown", fcf_before_fmls, FMT_DOLLAR, bold=True, total=True)

    # =========================================================================
    # 4. INTEREST SUMMARY
    # =========================================================================
    W.section_header("4.  INTEREST CALCULATION")
    W.year_header_row("Year 0")
    W.formula_row(None,"Revolver (incl. commitment fee)",
                  [None]+[f"={r('rev_interest',y)}" for y in range(1,N+1)], FMT_DOLLAR)
    W.formula_row(None,"Term Loan",
                  [None]+[f"={r('tl_interest',y)}"  for y in range(1,N+1)], FMT_DOLLAR)
    W.formula_row(None,"Mezzanine — Cash Interest",
                  [None]+[f"={r('mezz_cash_int',y)}" for y in range(1,N+1)], FMT_DOLLAR)
    W.formula_row(None,"Mezzanine — PIK Accretion",
                  [None]+[f"={r('mezz_pik',y)}"      for y in range(1,N+1)], FMT_DOLLAR)
    W.total_row(None,"Total Cash Interest Expense",
                [None]+[f"={r('rev_interest',y)}+{r('tl_interest',y)}+{r('mezz_cash_int',y)}"
                        for y in range(1,N+1)])
    W.blank()
    W.formula_row(None,"Amortization of Financing Fees",
                  [None]+[
                      f"=IF({y}<={i('FINANCING_FEE_AMORT_YEARS')},{W.txn_ref('annual_fee_amort')},0)"
                      for y in range(1,N+1)
                  ], FMT_DOLLAR, color=BLUE_FONT)
    W.total_row(None,"Total Int + Fee Amort",
                [None]+[
                    f"={r('rev_interest',y)}+{r('tl_interest',y)}+{r('mezz_cash_int',y)}"
                    f"+IF({y}<={i('FINANCING_FEE_AMORT_YEARS')},{W.txn_ref('annual_fee_amort')},0)"
                    for y in range(1,N+1)
                ])
    W.blank()

    # =========================================================================
    # 5. BALANCE SHEET SUMMARY
    # =========================================================================
    W.section_header("5.  BALANCE SHEET SUMMARY")
    W.year_header_row("Txn")

    bs_cash_fmls = [f"={i('CASH_TO_BS')}"] + [f"={r('cash_end',y)}" for y in range(1,N+1)]
    W.formula_row("bs_cash","Cash", bs_cash_fmls, FMT_DOLLAR)
    W.blank()

    bs_rev_fmls  = [f"=0"] + [f"={r('rev_end',y)}" for y in range(1,N+1)]
    bs_mezz_fmls = [f"={W.txn_ref('mezz_principal')}"] + [f"={r('mezz_end',y)}" for y in range(1,N+1)]
    bs_tl_fmls   = [f"={W.txn_ref('tl_principal')}"]   + [f"={r('tl_end',y)}"   for y in range(1,N+1)]
    W.formula_row("bs_rev", "Revolver",        bs_rev_fmls,  FMT_DOLLAR)
    W.formula_row("bs_mezz","Mezzanine (PIK)",  bs_mezz_fmls, FMT_DOLLAR)
    W.formula_row("bs_tl",  "Term Loan",        bs_tl_fmls,   FMT_DOLLAR)

    td_fmls = [f"={W.txn_ref('total_debt')}"] + [
        f"={r('bs_rev',y)}+{r('bs_mezz',y)}+{r('bs_tl',y)}" for y in range(1,N+1)
    ]
    W.total_row("bs_td","Total Debt", td_fmls)
    pct_fmls = [None] + [
        f"={r('bs_td',y)}/{W.txn_ref('total_debt')}" for y in range(1,N+1)
    ]
    W.formula_row(None,"% of Initial Debt", pct_fmls, FMT_PCT, indent=4)
    W.blank()

    nd_fmls = [
        f"={r('bs_td',y)}-{r('bs_cash',y)}" for y in range(N+1)
    ]
    W.total_row("bs_nd","Total Net Debt", nd_fmls)
    W.blank()

    lev_fmls = [f"={W.txn_ref('total_debt')}/{i('ENTRY_EBITDA')}"] + [
        f"={r('bs_td',y)}/{r('ebitda',y)}" for y in range(1,N+1)
    ]
    nlev_fmls = [f"={r('bs_nd',0)}/{i('ENTRY_EBITDA')}"] + [
        f"={r('bs_nd',y)}/{r('ebitda',y)}" for y in range(1,N+1)
    ]
    W.formula_row(None,"Total Debt / EBITDA",     lev_fmls,  FMT_MULT)
    W.formula_row(None,"Total Net Debt / EBITDA", nlev_fmls, FMT_MULT)
    W.blank()

    # =========================================================================
    # 6. RETURNS CALCULATION
    # =========================================================================
    W.section_header("6.  RETURNS CALCULATION")

    # Header with exit year labels
    r_hdr = W.row
    if not W.simulate:
        _apply(ws.cell(r_hdr, COL_LABEL, "  ($mm)"), h="left", bold=True, size=9, fill=SUB_BG)
        for y in range(1, N+1):
            _apply(ws.cell(r_hdr, _col(y-1), f"Exit Yr {y}"), bold=True, size=9, fill=SUB_BG)
    W.row += 1

    def ret_formula_row(label, fmls_by_year, fmt=FMT_DOLLAR, bold=False):
        rr = W.row
        if not W.simulate:
            _apply(ws.cell(rr, COL_LABEL, "  "+label), h="left", size=9, bold=bold)
            for y in range(1, N+1):
                c = ws.cell(rr, _col(y-1), fmls_by_year[y-1])
                _apply(c, fmt=fmt, size=9, bold=bold)
                if bold:
                    c.border = _top_border()
        W.row += 1

    ret_formula_row("LTM EBITDA",
                    [r('ebitda',y) for y in range(1,N+1)])
    ret_formula_row("(x) Exit Multiple",
                    [f"={i('EXIT_MULTIPLE')}"]*N, FMT_MULT)

    # write exit multiple row directly with colour
    if not W.simulate:
        for y in range(1,N+1):
            _apply(ws.cell(W.row-1, _col(y-1)), color=BLUE_FONT)

    ret_formula_row("Total Enterprise Value",
                    [f"={r('ebitda',y)}*{i('EXIT_MULTIPLE')}" for y in range(1,N+1)])
    ret_formula_row("Less: Net Debt",
                    [f"=-({r('bs_td',y)}-{r('bs_cash',y)})" for y in range(1,N+1)])
    ret_formula_row("Equity Value",
                    [f"={r('ebitda',y)}*{i('EXIT_MULTIPLE')}-({r('bs_td',y)}-{r('bs_cash',y)})"
                     for y in range(1,N+1)])
    ret_formula_row("(x) PE Ownership",
                    [f"=1-{i('ROLLOVER_EQUITY_PCT')}"]*N, FMT_PCT)
    ret_formula_row("Value to PE Firm",
                    [f"=({r('ebitda',y)}*{i('EXIT_MULTIPLE')}-({r('bs_td',y)}-{r('bs_cash',y)}))*(1-{i('ROLLOVER_EQUITY_PCT')})"
                     for y in range(1,N+1)],
                    FMT_DOLLAR, bold=True)
    W.blank()

    # IRR and MoM — using CAGR formula (exact for single entry/exit CF)
    for lbl, fmls, fmt in [
        ("MoM",
         [f"=(({r('ebitda',y)}*{i('EXIT_MULTIPLE')}-({r('bs_td',y)}-{r('bs_cash',y)}))*(1-{i('ROLLOVER_EQUITY_PCT')}))/{W.txn_ref('pe_invest')}"
          for y in range(1,N+1)],
         FMT_MOM),
        ("IRR",
         [f"=((({r('ebitda',y)}*{i('EXIT_MULTIPLE')}-({r('bs_td',y)}-{r('bs_cash',y)}))*(1-{i('ROLLOVER_EQUITY_PCT')}))/{W.txn_ref('pe_invest')})^(1/{y})-1"
          for y in range(1,N+1)],
         FMT_IRR),
    ]:
        rr = W.row
        if not W.simulate:
            _apply(ws.cell(rr, COL_LABEL, "  "+lbl), h="left", bold=True, size=9)
            for y in range(1,N+1):
                c = ws.cell(rr, _col(y-1), fmls[y-1])
                _apply(c, fmt=fmt, bold=True, size=9)
                c.border = _top_border()
        W.row += 1
    W.blank()

    # =========================================================================
    # 7. SENSITIVITY TABLES
    # =========================================================================
    W.section_header("7.  SENSITIVITY TABLES  (Terminal Year Exit)")

    import numpy as np
    entry_mults = np.round(np.arange(
        cfg.SENS_ENTRY_MIN, cfg.SENS_ENTRY_MAX + cfg.SENS_ENTRY_STEP/2, cfg.SENS_ENTRY_STEP), 4)
    exit_mults  = np.round(np.arange(
        cfg.SENS_EXIT_MIN,  cfg.SENS_EXIT_MAX  + cfg.SENS_EXIT_STEP/2,  cfg.SENS_EXIT_STEP), 4)

    # Terminal-year net debt reference (fixed across sensitivity)
    nd_ref   = f"({r('bs_td',N)}-{r('bs_cash',N)})"
    ebitda_n = r('ebitda', N)
    td_ref   = W.txn_ref('total_debt')
    fees_ref = W.txn_ref('total_fees')

    def _sens_table(title, fmt, is_irr: bool):
        W.sub_header(title)
        # Column headers (exit multiples)
        hdr_row = W.row
        if not W.simulate:
            _apply(ws.cell(hdr_row, COL_LABEL, "  Entry \\ Exit"), h="left", bold=True, size=9)
            for j, xm in enumerate(exit_mults):
                c = ws.cell(hdr_row, COL_LABEL + 1 + j, float(xm))
                _apply(c, fmt=FMT_MULT, bold=True, size=9, fill=SUB_BG)
        W.row += 1

        for i_idx, em in enumerate(entry_mults):
            base_row_flag = abs(float(em) - cfg.ENTRY_MULTIPLE) < 0.001
            data_row = W.row
            if not W.simulate:
                em_cell = ws.cell(data_row, COL_LABEL, f"  {em:.1f}x")
                _apply(em_cell, h="left", bold=base_row_flag, size=9)

            for j, xm in enumerate(exit_mults):
                base_cell_flag = base_row_flag and abs(float(xm) - cfg.EXIT_MULTIPLE) < 0.001
                col = COL_LABEL + 1 + j

                # em is a hardcoded float (label cell contains text "11.0x", not a number)
                # xm_ref references the header cell which stores a real float
                xm_ref  = f"${_cl(col)}${hdr_row}"
                em_val  = float(em)   # use numeric value directly in formula

                # pe_invest for this entry multiple:
                # = (ENTRY_EBITDA * em + CASH_TO_BS + total_fees + TX_EXPENSES - total_debt) * (1-ROLLOVER)
                pe_i = (f"=({i('ENTRY_EBITDA')}*{em_val}+{i('CASH_TO_BS')}+{fees_ref}"
                        f"+{i('TRANSACTION_EXPENSES')}-{td_ref})*(1-{i('ROLLOVER_EQUITY_PCT')})")

                # value to PE for this exit multiple:
                # = (terminal_ebitda * xm - terminal_net_debt) * (1-ROLLOVER)
                v2pe = (f"({ebitda_n}*{xm_ref}-{nd_ref})*(1-{i('ROLLOVER_EQUITY_PCT')})")

                if is_irr:
                    # IRR = (MoM)^(1/N) - 1
                    fml = (f"=({v2pe}/({i('ENTRY_EBITDA')}*{em_val}+{i('CASH_TO_BS')}"
                           f"+{fees_ref}+{i('TRANSACTION_EXPENSES')}-{td_ref}))^(1/{i('HOLD_YEARS')})-1")
                else:
                    fml = (f"={v2pe}/({i('ENTRY_EBITDA')}*{em_val}+{i('CASH_TO_BS')}"
                           f"+{fees_ref}+{i('TRANSACTION_EXPENSES')}-{td_ref})")

                if not W.simulate:
                    c = ws.cell(data_row, col, fml)
                    _apply(c, fmt=fmt, bold=base_cell_flag, size=9,
                           fill=SENS_HL if base_cell_flag else None)
            W.row += 1
        W.blank()

    _sens_table("IRR — Entry Multiple vs Exit Multiple", FMT_IRR, is_irr=True)
    _sens_table("MoM — Entry Multiple vs Exit Multiple", FMT_MOM, is_irr=False)


# =============================================================================
# Public entry point
# =============================================================================

def write_excel(
    txn: dict,
    results: dict,
    returns_data: list,
    sens: dict,
    overrides: dict | None = None,
    output_path: str | Path = "lbo_output.xlsx",
):
    overrides = overrides or {}

    wb = Workbook()
    wb.remove(wb.active)

    # ── Sheet 1: Inputs ───────────────────────────────────────────────────────
    inp_cells = _write_inputs_sheet(wb, overrides)

    # ── Sheet 2: Model ────────────────────────────────────────────────────────
    ws_model = wb.create_sheet("Model")
    ws_model.sheet_view.showGridLines = False
    N = cfg.HOLD_YEARS

    ws_model.column_dimensions[_cl(COL_LABEL)].width = LABEL_W
    for y in range(N + 1):
        ws_model.column_dimensions[_cl(_col(y))].width = DATA_W

    # ── Pass 1: simulation — count rows, populate reg[], write nothing ────────
    W_sim = _W(ws_model, N, simulate=True)
    _build_model_formulas(W_sim, inp_cells, N)

    # ── Pass 2: real write — pre-populated reg[] so all r() calls resolve ─────
    W = _W(ws_model, N, simulate=False)
    W.reg = W_sim.reg.copy()          # forward-references now pre-resolved
    _build_model_formulas(W, inp_cells, N)

    ws_model.freeze_panes = f"{_cl(COL_TXN)}5"

    output_path = Path(output_path)
    wb.save(output_path)
    print(f"\n  ✓  Excel model saved → {output_path.resolve()}")
    print("     Open the file and edit any blue cell on the Inputs sheet.")
    print("     The entire Model sheet recalculates automatically.")
    return output_path
